from download import download_file
from zipExtract import extract_file
from calculcation import cost_savings, calculations, imp_cost
from convertFormat import convert_xls_to_xlsx
from assessSheetExtraction import copy_rows_with_values, copy_first_rows, count_assem
from reccSheetExtraction import copy_rows_with_value, copy_first_row
from termCopy import copy_term
from datetime import datetime, timedelta
from arcCodes import arc_code_sheet
from labels import label_for_assem, label_for_recc
from formula_replace import pop_rows, replace_cell_value
from plots import plot_creation
import openpyxl
import xlwings as xw

#Universal Directory this is the path where all files will be saved ==============================================================================================
universal_dir = '/Users/maanitmalhan/Documents/IAC_Center/excel-data-iac/files/' 
#MAKE SURE TO USE THIS FORMAT AND END THE PATH WITH 'excel-data-iac/files/' OR THERE WILL BE ERRORS================================================================


#create SNE workbook
workbook = openpyxl.Workbook()
workbook.save(f'{universal_dir}/SNE_IAC_Database.xlsx')
print("File Created!")

#Download ZIP from server 
file_url = 'https://iac.university/storage/IAC_Database.zip'
destination_path = f'{universal_dir}IAC_Database.zip'

download_file(file_url, destination_path)
print(f"File downloaded successfully to {destination_path}")

#Extract file from ZIP
current_date = datetime.now()
yesterday = current_date - timedelta(days=1)    
date = current_date.strftime('%Y%m%d')

zip_file_path = f'{universal_dir}IAC_Database.zip'
file_to_extract = f'IAC_Database_{date}.xls'
extraction_path = universal_dir

extract_file(zip_file_path, file_to_extract, extraction_path)

print(f"File '{file_to_extract}' extracted successfully to '{extraction_path}'.")

#Convert Format from XLS to XLSX
input_path = f"{universal_dir}IAC_Database_{date}.xls"
output_path = f"{universal_dir}IAC_Database_{date}.xlsx"
    
convert_xls_to_xlsx(input_path, output_path)

#Extract Assesment data
source_workbook = openpyxl.load_workbook(f"{universal_dir}IAC_Database_{date}.xlsx")
source_sheet = source_workbook['ASSESS']  #name of your source sheet

destination_workbook = openpyxl.load_workbook(f'{universal_dir}SNE_IAC_Database.xlsx')
default_sheet = destination_workbook.active
default_sheet.title = "ASSESS"  
destination_sheet = destination_workbook['ASSESS']
target_value = 'UC'

copy_first_rows(source_sheet, destination_workbook, destination_sheet)
copy_rows_with_values(source_sheet, destination_workbook, destination_sheet, target_value)
count_assem(destination_sheet)
destination_workbook.save(f'{universal_dir}SNE_IAC_Database.xlsx')
label_for_assem()
print("Assessment data extracted successfully!")

#Extract Recommendation data
source_workbook = openpyxl.load_workbook(f"{universal_dir}IAC_Database_{date}.xlsx")
source_sheet = source_workbook['RECC5']  #name of your source sheet

destination_workbook = openpyxl.load_workbook(f"{universal_dir}SNE_IAC_Database.xlsx")
new_sheet = destination_workbook.create_sheet(title="RECC")
destination_sheet = destination_workbook['RECC']
target_column_index = 2  

copy_first_row(source_sheet, destination_workbook, destination_sheet)

for i in range(1, 10):
    target_value = 'UC230' + str(i)
    copy_rows_with_value(source_sheet, destination_workbook, destination_sheet, target_value, target_column_index)

for i in range(10, 100):
    target_value = 'UC23' + str(i)
    copy_rows_with_value(source_sheet, destination_workbook, destination_sheet, target_value, target_column_index)


destination_workbook.save(f"{universal_dir}SNE_IAC_Database.xlsx")
label_for_recc()

print("Recommendation data extracted successfully!")

#Term copy 
copy_term(date)
print("Term data copied successfully!")
#ARC codes
destination_workbook = openpyxl.load_workbook(f"{universal_dir}SNE_IAC_Database.xlsx")
arc_code_sheet(destination_workbook)
destination_workbook.save(f"{universal_dir}SNE_IAC_Database.xlsx")
print("ARC Codes imported successfully!")


destination_workbook = openpyxl.load_workbook(f"{universal_dir}SNE_IAC_Database.xlsx")
cost_savings(destination_workbook, 'K')
cost_savings(destination_workbook, 'O')
cost_savings(destination_workbook, 'S')
cost_savings(destination_workbook, 'W')
imp_cost(destination_workbook, 'G')
calculations(destination_workbook)
destination_workbook.save(f"{universal_dir}SNE_IAC_Database.xlsx")
print("Calculations done successfully!")

destination_workbook.create_sheet(title="Graphs")

destination_workbook.close()
source_workbook.close()

#formula replacement with calculated values
workbook = openpyxl.load_workbook(f"{universal_dir}SNE_IAC_Database.xlsx")
populated = pop_rows(workbook)


file_path = f"{universal_dir}SNE_IAC_Database.xlsx"
sheet_name = "RECC"


replace_cell_value(file_path, sheet_name, f"G{populated+2}")
replace_cell_value(file_path, sheet_name, f"G{populated+4}")
replace_cell_value(file_path, sheet_name, f"K{populated+2}")
replace_cell_value(file_path, sheet_name, f"O{populated+2}")
replace_cell_value(file_path, sheet_name, f"O{populated+4}")
replace_cell_value(file_path, sheet_name, f"S{populated+2}")
replace_cell_value(file_path, sheet_name, f"W{populated+2}")
replace_cell_value(file_path, sheet_name, f"W{populated+4}")


print('formula replaced with calculated values!')

workbook.close()
#Plot Creation
destination_workbook = openpyxl.load_workbook(f"{universal_dir}SNE_IAC_Database.xlsx", data_only=True)
plot_creation(destination_workbook)

destination_workbook.save(f"{universal_dir}SNE_IAC_Database.xlsx")
destination_workbook.close()

print("File Prepared! and saved as 'SNE_IAC_Database.xlsx' plots saved in file")
exit()