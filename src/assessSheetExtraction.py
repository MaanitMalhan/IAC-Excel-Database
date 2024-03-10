import openpyxl
from copy import copy
from datetime import datetime, timedelta

current_date = datetime.now()
yesterday = current_date - timedelta(days=1)    
date = yesterday.strftime('%Y%m%d')


def copy_first_rows(source_sheet, target_workbook, target_sheet):
    for source_cell in source_sheet[1]:
        target_sheet.cell(row=1, column=source_cell.column, value=source_cell.value)


def copy_rows_with_values(source_sheet, target_workbook, target_sheet, target_value):
    for row in source_sheet.iter_rows():
        for cell in row:
            if cell.value == target_value:
                # Create a new row in the target sheet
                target_row = target_sheet.max_row + 1
                # Copy each cell from the source row to the target row
                for source_cell in row:
                    target_sheet.cell(row=target_row, column=source_cell.column, value=source_cell.value)

def count_assem(target_sheet):
    count = 0
    column_to_check = 'A'
    populated_rows = 0
    # Iterate over rows in the column and count populated ones
    for row in target_sheet.iter_rows(min_row=1, max_row=target_sheet.max_row, min_col=1, max_col=1):
        cell_value = row[0].value
        if cell_value is not None and str(cell_value).strip() != '':
            populated_rows += 1
    populated_rows -= 1
    return populated_rows



# Specify the value you want to find and copy
target_value = 'UC'
