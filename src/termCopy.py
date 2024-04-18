import openpyxl
from datetime import datetime, timedelta
from app import *

def copy_term(date):
# Load the source workbook
    source_workbook = openpyxl.load_workbook(f"{universal_dir}IAC_Database_{date}.xlsx")
    source_sheet = source_workbook['Terms']  

# Load the destination workbook
    destination_workbook = openpyxl.load_workbook(f'{universal_dir}SNE_IAC_Database.xlsx')
    new_sheet = destination_workbook.create_sheet(title="Terms")
    destination_sheet = destination_workbook['Terms']  

    for row in source_sheet.iter_rows(min_row=1, max_row=source_sheet.max_row, min_col=1, max_col=source_sheet.max_column):
        for cell in row:
            destination_sheet[cell.coordinate].value = cell.value
   
    img = openpyxl.drawing.image.Image(f'{universal_dir}IAC_PREF_APRIL_10_2024.png')
    img.anchor = 'K1'
    destination_sheet.add_image(img)

    destination_sheet['K60'] = 'In Image SNE IAC Data as of APRIL 10, 2024'

# Save the changes to the destination workbook
    destination_workbook.save(f'{universal_dir}SNE_IAC_Database.xlsx')
