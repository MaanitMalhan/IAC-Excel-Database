import openpyxl
from assessSheetExtraction import count_assem
from reccSheetExtraction import count_recc
from openpyxl.styles import Font
#Add recommended savings, plant energy costs and number of reccs





def cost_savings(workbook, column):
    destination_workbook = workbook
    destination_sheet = destination_workbook['RECC']
    #Add recommended savings, plant energy costs and number of reccs
    count = 0
    column_to_check = column
    populated_rows = 0
    # Iterate over rows in the column and count populated ones
    for row in destination_sheet.iter_rows(min_row=1, max_row=destination_sheet.max_row, min_col=1, max_col=1):
        cell_value = row[0].value
        if cell_value is not None and str(cell_value).strip() != '':
            populated_rows += 1

    destination_sheet[f'{column_to_check}{populated_rows+1}'] = 'Total_type_recommended_savings'
    destination_sheet[f'{column_to_check}{populated_rows+1}'].font = Font(bold=True)
    destination_sheet[f'{column_to_check}{populated_rows+2}'] = f'=SUM({column_to_check}2:{column_to_check}{populated_rows})'
    if column == 'W':
        destination_sheet[f'{column_to_check}{populated_rows+3}'] = 'Average_Savings_From_Recommendation'
        destination_sheet[f'{column_to_check}{populated_rows+3}'].font = Font(bold=True)
        destination_sheet[f'{column_to_check}{populated_rows+4}'] = f'=AVERAGE(K{populated_rows+2},O{populated_rows+2},S{populated_rows+2},W{populated_rows+2})'
    if column == 'O':
        destination_sheet[f'{column_to_check}{populated_rows+3}'] = 'Absolute_Savings_From_Recommendation'
        destination_sheet[f'{column_to_check}{populated_rows+3}'].font = Font(bold=True)
        destination_sheet[f'{column_to_check}{populated_rows+4}'] = f'=SUM(K{populated_rows+2},O{populated_rows+2},S{populated_rows+2},W{populated_rows+2})'
    
    return populated_rows+2

def imp_cost(workbook, column):
    destination_workbook = workbook
    destination_sheet = destination_workbook['RECC']
    #Add recommended savings, plant energy costs and number of reccs
    count = 0
    column_to_check = column
    populated_rows = 0
    # Iterate over rows in the column and count populated ones
    for row in destination_sheet.iter_rows(min_row=1, max_row=destination_sheet.max_row, min_col=1, max_col=1):
        cell_value = row[0].value
        if cell_value is not None and str(cell_value).strip() != '':
            populated_rows += 1

    destination_sheet[f'{column_to_check}{populated_rows+1}'] = 'Total_Implementation_Cost'
    destination_sheet[f'{column_to_check}{populated_rows+1}'].font = Font(bold=True)
    destination_sheet[f'{column_to_check}{populated_rows+2}'] = f'=SUM({column_to_check}2:{column_to_check}{populated_rows})'
    if column == 'G':
        destination_sheet[f'{column_to_check}{populated_rows+3}'] = 'Average_Implementation_Cost'
        destination_sheet[f'{column_to_check}{populated_rows+3}'].font = Font(bold=True)
        destination_sheet[f'{column_to_check}{populated_rows+4}'] = f'=G{populated_rows+2}/{populated_rows}'
    return populated_rows+2




def calculations(workbook):
    destination_workbook = workbook
    calc_sheet = destination_workbook.create_sheet(title="calculation")
    destination_sheet = destination_workbook['calculation']
    recc_sheet = destination_workbook['RECC']
    #rec_location = cost_savings(destination_workbook)

    #Add recommended savings, plant energy costs and number of reccs
    destination_sheet['A1'] = 'Total_number_of_assessments'
    destination_sheet['B1'] = 'Total_number_of_recommendations'
    destination_sheet['C1'] = 'Average_number_of_recommendations_per_assessment'

    
    #Add recommended savings, plant energy costs and number of reccs
    destination_sheet['A2'] = count_assem(destination_workbook['ASSESS'])
    destination_sheet['B2'] = count_recc(destination_workbook['RECC'])

    destination_sheet['C2'] = destination_sheet['B2'].value / destination_sheet['A2'].value

