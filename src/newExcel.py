import openpyxl
import os
from datetime import datetime, timedelta

current_date = datetime.now()
yesterday = current_date - timedelta(days=1)
date = yesterday.strftime('%Y%m%d')

iac_workbook = openpyxl.load_workbook(f'/Users/maanitmalhan/Documents/IAC Center/excel-data-iac/IAC_Database_{date}.xls')
iac_assess_worksheet = iac_workbook["ASSESS"]

sne_workbook = openpyxl.Workbook()
sne_worksheet = sne_workbook.active
sne_worksheet.title = "CopiedRows"

sne_workbook.save('/Users/maanitmalhan/Documents/IAC Center/excel-data-iac/SNE_IAC_Database.xls')


def copy_rows_containing_value(source_file, source_sheet, column_name, search_value, destination_file, destination_sheet):
    try:
        # Open the source workbook
        source_wb = source_file

        # Open the source worksheet
        source_ws = source_sheet

        # Find the column index based on the column name
        column_index = None
        for col_num, col in enumerate(source_ws.iter_cols(min_row=1, max_row=1, values_only=True), start=1):
            if col[0] == column_name:
                column_index = col_num
                break

        if column_index is None:
            raise ValueError(f"Column '{column_name}' not found in '{source_sheet}' of '{source_file}'")

        # Initialize a list to store rows containing the search value
        matching_rows = []

        # Iterate through the rows and find matching values
        for row in source_ws.iter_rows(min_row=2, values_only=True):
            if row[column_index - 1] and search_value in str(row[column_index - 1]).upper():
                matching_rows.append(row)

        # Create or open the destination workbook
        try:
            destination_wb = destination_file
        except FileNotFoundError:
            print(f"File '{destination_file}' not found.")

        # Create a new worksheet in the destination workbook
        destination_ws = destination_wb.create_sheet(title=destination_sheet)

        # Write matching rows to the new worksheet
        for row_data in matching_rows:
            destination_ws.append(row_data)

        # Save the destination workbook
        destination_wb.save(destination_file)

        print(f"Rows containing '{search_value}' in '{column_name}' copied from '{source_sheet}' in '{source_file}' "
              f"to '{destination_sheet}' in '{destination_file}'")

    except Exception as e:
        print(f"An error occurred: {str(e)}")

# Example usage:
source_file = iac_workbook
source_sheet = 'ASSESS'
column_name_to_search = "B"  # Replace with the actual column name
search_value = 'UC'
destination_file = sne_workbook
destination_sheet = 'CopiedRows'

